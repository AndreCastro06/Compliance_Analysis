import pandas as pd
import os
from openpyxl import load_workbook

def processar_meds(caminho_csv: str, caminho_excel_saida: str):
    # Dados para identificação
    coluna_cnpj = 'CpfCnpjCreditado'
    coluna_id_med = 'IdNotificacaoInfracao'

    # Carregar MEDs novos (Separador do csv: ponto e vírgula)
    df = pd.read_csv(caminho_csv, sep=';', dtype=str)

    # Considerar apenas registros com status "Recebida"
    df = df[df['Fluxo'].str.strip().str.lower() == 'recebida']

    # Lê lista de CNPJs Pix Indiretos
    cnpjs_excluidos = set()
    df_Pix_Indiretos = pd.DataFrame()

    try:
        df_excluir = pd.read_csv("data/excluir_cnpjs.csv", dtype=str)
        cnpjs_excluidos = set(df_excluir["CNPJ"].str.strip())

        # Separa os MEDs a serem excluídos
        df_Pix_Indiretos = df[df[coluna_cnpj].isin(cnpjs_excluidos)]

        # Salva os excluídos em um arquivo separado
        if not df_Pix_Indiretos.empty:
           df_Pix_Indiretos.to_excel("output/MEDS_Pix_Indiretos.xlsx", sheet_name="Pix_Indiretos", index=False)

    except FileNotFoundError:
        print("⚠️ Arquivo 'excluir_cnpjs.csv' não encontrado. Nenhum CNPJ será excluído.")


    # Separa MEDs sem CNPJ creditado
    df_sem_cnpj = df[df[coluna_cnpj].isna() | (df[coluna_cnpj].str.strip() == '')]

    from meds_analyzer.atualizar_meds_sem_cnpj import atualizar_meds_sem_cnpj

    atualizar_meds_sem_cnpj(
        df_sem_cnpj=df_sem_cnpj,
        coluna_id_med=coluna_id_med,
        output_path="output/MEDS_SEM_CNPJ.xlsx"
    )

    # Remove do DataFrame principal para não tentar criar aba sem nome
    df = df[~(df[coluna_cnpj].isna() | (df[coluna_cnpj].str.strip() == ''))]
        # Remove os excluídos do DataFrame principal
    df = df[~df[coluna_cnpj].isin(cnpjs_excluidos)]

    # Gera arquivo final com os MEDs válidos por empresa
    if os.path.exists(caminho_excel_saida):
        writer = pd.ExcelWriter(caminho_excel_saida, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        workbook = load_workbook(caminho_excel_saida)

        for cnpj, grupo in df.groupby(coluna_cnpj):
            nome_aba = cnpj  # Usa CNPJ completo como nome da aba
            try:
                if nome_aba in workbook.sheetnames:
                    df_existente = pd.read_excel(caminho_excel_saida, sheet_name=nome_aba, dtype=str)
                    df_combinado = pd.concat([df_existente, grupo], ignore_index=True)
                    df_combinado.drop_duplicates(subset=[coluna_id_med], inplace=True)
                else:
                    df_combinado = grupo
            except Exception as e:
                print(f"Erro ao processar a aba {nome_aba}: {e}")
                df_combinado = grupo

            df_combinado.to_excel(writer, sheet_name=nome_aba, index=False)
        writer.close()
    else:
        with pd.ExcelWriter(caminho_excel_saida, engine='openpyxl') as writer:
            for cnpj, grupo in df.groupby(coluna_cnpj):
                nome_aba = cnpj
                grupo.to_excel(writer, sheet_name=nome_aba, index=False)