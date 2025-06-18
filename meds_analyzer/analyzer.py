import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
import unicodedata
import re

def formatar_cabecalhos(path_excel):
    wb = load_workbook(path_excel)

    fill_azul = PatternFill(start_color='FF010AFD', end_color='FF010AFD', fill_type='solid')
    fonte_branca_negrito = Font(color='FFFFFFFF', bold=True)
    alinhamento_central = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fonte_vermelha_negrito = Font(color='FF0000', bold=True)

    for aba in wb.sheetnames:
        ws = wb[aba]

        for cell in ws[1]:
            cell.fill = fill_azul
            cell.font = fonte_branca_negrito
            cell.alignment = alinhamento_central

        headers = [cell.value for cell in ws[1]]
        try:
            col_med_amount = headers.index('% Valor MEDs x Pix-In (Dia)') + 1
            col_acumulado = headers.index('% Valor MEDs x Pix-In (Semana)') + 1
        except ValueError:
            continue  # pula a aba se não encontrar as colunas necessárias

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try:
                valor_raw_1 = row[col_med_amount - 1].value
                if isinstance(valor_raw_1, str):
                    valor_1 = float(re.sub(r'[^\d.,]', '', valor_raw_1).replace(',', '.')) / 100 if '%' in valor_raw_1 else float(valor_raw_1.replace(',', '.'))
                else:
                    valor_1 = float(valor_raw_1)
                if valor_1 > 0.0035:
                    row[col_med_amount - 1].font = fonte_vermelha_negrito
            except:
             pass

        try:
            valor_raw_2 = row[col_acumulado - 1].value
            if isinstance(valor_raw_2, str):
                valor_2 = float(re.sub(r'[^\d.,]', '', valor_raw_2).replace(',', '.')) / 100 if '%' in valor_raw_2 else float(valor_raw_2.replace(',', '.'))
            else:
                valor_2 = float(valor_raw_2)
            if valor_2 > 0.0035:
                row[col_acumulado - 1].font = fonte_vermelha_negrito
        except:

            try:
                valor_raw_2 = row[col_acumulado - 1].value
                if isinstance(valor_raw_2, str):
                    valor_2 = float(re.sub(r'[^\d.,]', '', valor_raw_2).replace(',', '.')) / 100 if '%' in valor_raw_2 else float(valor_raw_2.replace(',', '.'))
                else:
                    valor_2 = float(valor_raw_2)
                if valor_2 > 0.0035:
                    row[col_acumulado - 1].font = fonte_vermelha_negrito
            except:
                pass

    wb.save(path_excel)
    print(f" Estilização aplicada: {path_excel}")

# Cnverter valores monetários no padrão brasileiro
def converter_valor_brasileiro(valor):
    if pd.isna(valor):
        return 0.0
    valor_str = str(valor).strip()
    valor_str = re.sub(r'[^\d,.-]', '', valor_str)  # remove caracteres estranhos
    if ',' in valor_str and '.' in valor_str:
        # Ex: 1.234,56 → 1234.56
        valor_str = valor_str.replace('.', '').replace(',', '.')
    elif ',' in valor_str:
        # Ex: 1234,56 → 1234.56
        valor_str = valor_str.replace(',', '.')
    return float(valor_str)


def converter_valor_brasileiro(valor):
    if pd.isna(valor):
        return 0.0
    valor_str = str(valor).strip()
    valor_str = re.sub(r'[^\d,.-]', '', valor_str)  # remove R$, espaços, etc.
    if ',' in valor_str and '.' in valor_str:
        # Caso típico brasileiro: '1.234.567,89'
        valor_str = valor_str.replace('.', '').replace(',', '.')
    elif ',' in valor_str:
        # Caso: '1234567,89'
        valor_str = valor_str.replace(',', '.')
    return float(valor_str)

def formatar_reais(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "R$ 0,00"


def normalizar_datas(dataframe, coluna):
    dataframe[coluna] = pd.to_datetime(dataframe[coluna]).dt.date
    return dataframe

def inserir_percentual_acumulado_mensal(df_analise):
    # Garantir datetime na coluna 'Data'
    df_analise["Data"] = pd.to_datetime(df_analise["Data"])
    df_analise["Ano"] = df_analise["Data"].dt.year
    df_analise["Mes"] = df_analise["Data"].dt.month

    # Converter colunas para número
    df_analise["Transacoes_Convertido"] = pd.to_numeric(df_analise["Qtd. Transações"], errors='coerce').fillna(0)
    df_analise["MEDs_Convertido"] = pd.to_numeric(df_analise["Qtd. MEDs"], errors='coerce').fillna(0)

    # Acumulados mês a mês
    df_analise["Acum_PIX"] = df_analise.groupby(["Ano", "Mes"])["Transacoes_Convertido"].cumsum()
    df_analise["Acum_MEDs"] = df_analise.groupby(["Ano", "Mes"])["MEDs_Convertido"].cumsum()

    # Percentual acumulado do mês
    df_analise["Percentual Abertura MEDs x Transações - Mês"] = df_analise.apply(
        lambda row: f"{(row['Acum_MEDs'] / row['Acum_PIX'] * 100):.2f}%" if row["Acum_PIX"] != 0 else "0,00%",
        axis=1
    )

    # Limpeza
    df_analise.drop(columns=[
        "Ano", "Mes", "Transacoes_Convertido", "MEDs_Convertido",
        "Acum_PIX", "Acum_MEDs"
    ], inplace=True)

    return df_analise

# Função de normalização de texto para analise de Golpes
palavras_chave = [
    "fraude", "triangulacao", "roubo", "furto", "golpe",
    "estelionato", "ameaca"
]

def normalizar(texto):
    texto = str(texto).lower()
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')  # Remove acentos
    texto = re.sub(r'[^\w\s]', '', texto)  # Remove pontuação
    return texto

def gerar_analise_geral_e_diaria( caminho_meds="data/meds.csv", caminho_transacional="data/metabase_píx", caminho_clientes="SEU_CAMINHO/Base Clientes - Microcash.xlsx", saida_geral="output/ MEDS - Analise Geral.xlsx",
    saida_diaria="output/MEDS - Report Diário.xlsx", intervalo_dias=3 ):
    abas_meds = pd.ExcelFile(caminho_meds).sheet_names
    abas_pix = pd.ExcelFile(caminho_transacional).sheet_names

    df_clientes = pd.read_excel(caminho_clientes, dtype=str)
    df_clientes["Documento"] = df_clientes["Documento"].str.replace(r'\D', '', regex=True)
    mapa_cnpj_para_nome = dict(zip(df_clientes["Documento"], df_clientes["Pessoa Nome"]))

    # DEBUG: Verifique se os CNPJs estão batendo
    print("🔍 Mapa CNPJ → Nome (amostra):")
    for k, v in list(mapa_cnpj_para_nome.items())[:5]:
        print(f"{k} → {v}")

    todas_analises = []
    for aba in abas_meds:
        cnpj_atual = aba.strip()
        nome_empresa = mapa_cnpj_para_nome.get(cnpj_atual, f"Empresa_{cnpj_atual}")[:31]
        try:
            df_meds = pd.read_excel(caminho_meds, sheet_name=aba, dtype=str)
            df_meds["ValorTransacao"] = pd.to_numeric(df_meds["ValorTransacao"].str.replace(',', '.'), errors='coerce').fillna(0)
            df_meds["Created At: Day"] = pd.to_datetime(df_meds["DtHrCriacaoNotificacaoInfracao"]).dt.date
        
            
        except Exception as e:
            print(f"Erro ao ler MEDs da aba {aba}: {e}")
            continue

        aba_transacional = cnpj_atual if cnpj_atual in abas_pix else None

        if aba_transacional:
            try:
                df_pix = pd.read_excel(caminho_transacional, sheet_name=aba_transacional, dtype=str)
                df_pix["Transactions"] = pd.to_numeric(df_pix["Transactions"].str.replace('.', '').str.replace(',', '.'), errors='coerce').fillna(0).astype(int)
                df_pix["Sum of Amount"] = df_pix["Sum of Amount"].apply(converter_valor_brasileiro).fillna(0).round(2)
                df_pix["Sum of Amount"] = df_pix["Sum of Amount"].apply(converter_valor_brasileiro).fillna(0).round(2)
                df_pix["Data"] = pd.to_datetime(df_pix["Created At: Day"]).dt.date
            except Exception as e:
                print(f"Erro ao ler transações da aba {aba_transacional}: {e}")
                df_pix = pd.DataFrame(columns=["Data", "Transactions", "Sum of Amount"])
        else:
            print(f"⚠️ Nenhuma aba de transações para {aba}. A análise será feita apenas com os MEDs.")
            df_pix = pd.DataFrame(columns=["Data", "Transactions", "Sum of Amount"])

        analise_diaria = []
        datas_meds = set(df_meds["Created At: Day"].unique())
        datas_pix = set(df_pix["Data"].unique()) if not df_pix.empty else set()
        todas_datas = sorted(datas_meds.union(datas_pix))
        for data in todas_datas:
            data_date = pd.to_datetime(data).date()
            meds_dia = df_meds[df_meds["Created At: Day"] == data_date]
            pix_dia = df_pix[df_pix["Data"] == data_date] if not df_pix.empty else pd.DataFrame(columns=["Transactions", "Sum of Amount"])
            total_meds = meds_dia["ValorTransacao"].sum()
            qtd_meds = len(meds_dia)
            meds_menor_500 = meds_dia[meds_dia["ValorTransacao"] < 500]
            meds_maior_500 = meds_dia[meds_dia["ValorTransacao"] >= 500]
            palavras_chave = ["fraude", "triangulação", "triangulacao", "roubo", "furto", "golpe", "estelionato","Triangulação","Golpe","Estelionato","Ameaça"]
            detalhes_series = meds_dia["DetalhesNotificacaoInfracao"].astype(str).apply(normalizar)
            qtd_suspeitas = detalhes_series.apply(lambda texto: any(p in texto for p in palavras_chave)).sum()

            total_pix = pix_dia["Sum of Amount"].sum()
            qtd_pix = pix_dia["Transactions"].sum()

            meds_acumulado = df_meds[(df_meds["Created At: Day"] >= data_date - timedelta(days=6)) & (df_meds["Created At: Day"] <= data_date)]["ValorTransacao"].sum()
            pix_acumulado = df_pix[(df_pix["Data"] >= data_date - timedelta(days=6)) & (df_pix["Data"] <= data_date)]["Sum of Amount"].sum()

            acumulado_percentual = f"{(meds_acumulado / pix_acumulado * 100):.2f}%" if pix_acumulado else "0,00%"

            # ✅ NOVO BLOCO PARA ACUMULADO MENSAL
            primeiro_dia_mes = data_date.replace(day=1)
            df_meds_mes = df_meds[(df_meds["Created At: Day"] >= primeiro_dia_mes) & (df_meds["Created At: Day"] <= data_date)]
            df_pix_mes = df_pix[(df_pix["Data"] >= primeiro_dia_mes) & (df_pix["Data"] <= data_date)]
            valor_meds_mes = df_meds_mes["ValorTransacao"].sum()
            valor_pix_mes = df_pix_mes["Sum of Amount"].sum()
            percentual_acumulado_mes_valor = f"{(valor_meds_mes / valor_pix_mes * 100):.2f}%" if valor_pix_mes else "0,00%"

            linha = {
                    "Data": data_date,
                    "Qtd. Transações": int(qtd_pix),
                    "Valor Pix-In": formatar_reais(total_pix),
                    "Qtd. MEDs": qtd_meds,
                    "Valor MEDs": formatar_reais(total_meds),
                    "MEDs < R$ 500": len(meds_menor_500),
                    "% MEDs < R$ 500": f"{(meds_menor_500['ValorTransacao'].sum() / total_meds * 100):.2f}%" if total_meds else "0,00%",
                    "MEDs >= R$ 500": len(meds_maior_500),
                    "% MEDs >= R$ 500": f"{(len(meds_maior_500) / qtd_meds * 100):.2f}%" if qtd_meds else "0,00%",
                    "% MEDs x Pix-In (Qtd)": f"{(qtd_meds / qtd_pix * 100):.2f}%" if qtd_pix else "0,00%",
                    "% Valor MEDs x Pix-In (Dia)": f"{(total_meds / total_pix * 100):.2f}%" if total_pix else "0,00%",
                    "% Valor MEDs x Pix-In (Semana)": acumulado_percentual,
                    "% Valor MEDs x Pix-In (Mês)": percentual_acumulado_mes_valor,
                    "QNT Suspeitas por Palavra-chave": qtd_suspeitas,
                    "% MEDs Suspeitas": f"{(qtd_suspeitas / qtd_meds * 100):.2f}%" if qtd_meds else "0,00%",
                                        }
            analise_diaria.append(linha)

        df_analise = pd.DataFrame(analise_diaria)
        df_analise = inserir_percentual_acumulado_mensal(df_analise)

        # Reorganiza colunas
        ordem_colunas = [
                        "Data",
                        "Qtd. Transações",
                        "Valor Pix-In",
                        "Qtd. MEDs",
                        "Valor MEDs",
                        "MEDs < R$ 500",
                        "% MEDs < R$ 500",
                        "MEDs >= R$ 500",
                        "% MEDs >= R$ 500",
                        "% MEDs x Pix-In (Qtd)",
                        "% Valor MEDs x Pix-In (Dia)",
                        "% Valor MEDs x Pix-In (Semana)",
                        "% Valor MEDs x Pix-In (Mês)",
                        "QNT Suspeitas por Palavra-chave",
                        "% MEDs Suspeitas"
            ]
        ordem_existente = [col for col in ordem_colunas if col in df_analise.columns]
        df_analise = df_analise[ordem_existente]
        todas_analises.append((nome_empresa, df_analise))

    # Garante os dados com linhas zeradas mesmo que sem MEDs ou PIX
    datas_completas = set()
    for _, df_empresa in todas_analises:
        datas_completas.update(df_empresa["Data"].unique())

    if not datas_completas:
        datas_completas = {datetime.today().date()}

    for i, (nome_empresa, df_empresa) in enumerate(todas_analises):
        datas_existentes = set(df_empresa["Data"].unique())
        datas_faltando = datas_completas - datas_existentes

        linhas_zeradas = []
        for data_faltante in sorted(datas_faltando):
            linhas_zeradas.append({
              "Data": data_faltante,
              "Qtd. Transações": 0,
              "Valor Pix-In": 0,
              "Qtd. MEDs": 0,
              "Valor MEDs": 0,
              "MEDs < R$ 500": 0,
              "% MEDs < R$ 500": "0,00%",
              "MEDs >= R$ 500": 0,
              "% MEDs >= R$ 500": "0,00%",
              "% MEDs x Pix-In (Qtd)": "0,00%",
              "% Valor MEDs x Pix-In (Dia)": "0,00%",
              "% Valor MEDs x Pix-In (Semana)": "0,00%",
              "% Valor MEDs x Pix-In (Mês)": "0,00%",
              "QNT Suspeitas por Palavra-chave": "0",
              "% MEDs Suspeitas": "0",
            })
        if linhas_zeradas:
            todas_analises[i] = (nome_empresa, pd.concat([df_empresa, pd.DataFrame(linhas_zeradas)], ignore_index=True))

    # Grava o relatório geral considerando se ja foi feito
    dados_existentes = {}
    if os.path.exists(saida_geral):
        print("📂 Carregando dados existentes do arquivo geral...")
        with pd.ExcelFile(saida_geral) as reader:
            for aba in reader.sheet_names:
                df_antigo = pd.read_excel(reader, sheet_name=aba, dtype=str)
                df_antigo["Data"] = pd.to_datetime(df_antigo["Data"]).dt.date
                dados_existentes[aba] = df_antigo

    with pd.ExcelWriter(saida_geral, engine='openpyxl', mode='w') as writer:
        for nome_empresa, df_novo in todas_analises:
            if nome_empresa in dados_existentes:
                df_antigo = dados_existentes[nome_empresa]
                df_novo["Data"] = pd.to_datetime(df_novo["Data"]).dt.date
                datas_antigas = set(df_antigo["Data"])
                df_novos_dias = df_novo[~df_novo["Data"].isin(datas_antigas)]
                df_combinado = pd.concat([df_antigo, df_novos_dias], ignore_index=True)
                df_combinado = df_combinado.sort_values("Data")
            else:
                df_combinado = df_novo.sort_values("Data")
            
            df_combinado["Data"] = pd.to_datetime(df_combinado["Data"]).dt.strftime("%d/%m/%Y")

            df_combinado.to_excel(writer, sheet_name=nome_empresa, index=False)

    # Pergunta e gera relatório diário

        # Solicita intervalo ao usuário
    data_inicio_str = input("📆 Digite a data de início da análise (dd/mm/aaaa): ")
    data_fim_str = input("📆 Digite a data de fim da análise (dd/mm/aaaa): ")

    try:
        data_inicio = datetime.strptime(data_inicio_str, "%d/%m/%Y").date()
        data_fim = datetime.strptime(data_fim_str, "%d/%m/%Y").date()
    except ValueError:
        print("❌ Formato de data inválido. Use dd/mm/aaaa.")
        return

    todas_datas_disponiveis = set()
    for _, df_empresa in todas_analises:
        todas_datas_disponiveis.update(df_empresa["Data"].unique())

    if not todas_datas_disponiveis:
        print("⚠️ Nenhuma data encontrada nos dados.")
        return
    
    # Normaliza tudo como date
    todas_datas_disponiveis = {pd.to_datetime(d).date() for d in todas_datas_disponiveis}

    primeira_data_disponivel = min(todas_datas_disponiveis)
    ultima_data_disponivel = max(todas_datas_disponiveis)

    print(f"\n📅 Intervalo de dados disponível: {primeira_data_disponivel.strftime('%d/%m/%Y')} até {ultima_data_disponivel.strftime('%d/%m/%Y')}")

    # Verifica se datas inseridas estão dentro do intervalo
    if data_inicio < primeira_data_disponivel:
        print(f"⚠️ Data de início ajustada para a primeira data disponível: {primeira_data_disponivel.strftime('%d/%m/%Y')}")
        data_inicio = primeira_data_disponivel.date

    if data_fim > ultima_data_disponivel:
        print(f"⚠️ Data de fim ajustada para a última data disponível: {ultima_data_disponivel.strftime('%d/%m/%Y')}")
        data_fim = ultima_data_disponivel

    # Gera o intervalo
    modo_consolidado = (data_fim - data_inicio).days > 0

    dias = pd.date_range(start=data_inicio, end=data_fim).date

    # Filtra as análises com base nos dias
    analises_diarias = []

    for nome_empresa, df in todas_analises:
        df["Data"] = pd.to_datetime(df["Data"]).dt.date
        df_filtrado = df[df["Data"].isin(dias)]
        if not df_filtrado.empty:
            analises_diarias.append((nome_empresa, df_filtrado))

    if analises_diarias:
        df_diario_unico = pd.DataFrame()

        for nome_empresa, df_filtrado in analises_diarias:
            df_filtrado["Cliente"] = nome_empresa
            if modo_consolidado:
                df_filtrado["Qtd. Transações"] = pd.to_numeric(df_filtrado["Qtd. Transações"], errors='coerce').fillna(0)
                df_filtrado["Qtd. MEDs"] = pd.to_numeric(df_filtrado["Qtd. MEDs"], errors='coerce').fillna(0)
                df_filtrado["MEDs < R$ 500"] = pd.to_numeric(df_filtrado["MEDs < R$ 500"], errors='coerce').fillna(0)
                df_filtrado["MEDs >= R$ 500"] = pd.to_numeric(df_filtrado["MEDs >= R$ 500"], errors='coerce').fillna(0)

                df_filtrado["Valor Pix-In"] = df_filtrado["Valor Pix-In"].apply(converter_valor_brasileiro)
                df_filtrado["Valor MEDs"] = df_filtrado["Valor MEDs"].apply(converter_valor_brasileiro)

                agrupado = df_filtrado.groupby("Cliente").agg({
                    "Qtd. Transações": "sum",
                    "Valor Pix-In": "sum",
                    "Qtd. MEDs": "sum",
                    "Valor MEDs": "sum",
                    "MEDs < R$ 500": "sum",
                    "MEDs >= R$ 500": "sum"
                }).reset_index()

                agrupado["% < R$ 500"] = agrupado["MEDs < R$ 500"] / agrupado["Valor MEDs"] * 100
                agrupado["% > R$ 500"] = agrupado["MEDs >= R$ 500"] / agrupado["Qtd. MEDs"] * 100
                agrupado["% Transaction"] = agrupado["Qtd. MEDs"] / agrupado["Qtd. Transações"] * 100
                agrupado["% Amount"] = agrupado["Valor MEDs"] / agrupado["Valor Pix-In"] * 100

                agrupado["Valor Pix-In"] = agrupado["Valor Pix-In"].apply(formatar_reais)
                agrupado["Valor MEDs"] = agrupado["Valor MEDs"].apply(formatar_reais)

                agrupado["% < R$ 500"] = agrupado["% < R$ 500"].apply(lambda x: f"{x:.2f}%")
                agrupado["% > R$ 500"] = agrupado["% > R$ 500"].apply(lambda x: f"{x:.2f}%")
                agrupado["% Transaction"] = agrupado["% Transaction"].apply(lambda x: f"{x:.2f}%")
                agrupado["% Amount"] = agrupado["% Amount"].apply(lambda x: f"{x:.2f}%")

                agrupado["Data"] = f"{data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
                agrupado["Ação"] = ""

                agrupado = agrupado[[
                    "Data", "Cliente", "Qtd. Transações", "Valor Pix-In", "Qtd. MEDs", "Valor MEDs",
                    "MEDs < R$ 500", "% < R$ 500", "MEDs >= R$ 500", "% > R$ 500",
                    "% Transaction", "% Amount", "Ação"
                ]]

                df_diario_unico = pd.concat([df_diario_unico, agrupado], ignore_index=True)

            else:
                df_filtrado["Data"] = pd.to_datetime(df_filtrado["Data"]).dt.strftime("%d/%m/%Y")
                df_filtrado["Ação"] = ""
                df_filtrado = df_filtrado[[
                    "Data", "Cliente", "Qtd. Transações", "Valor Pix-In", "Qtd. MEDs", "Valor MEDs",
                    "MEDs < R$ 500", "% MEDs < R$ 500", "MEDs >= R$ 500", "% MEDs >= R$ 500",
                    "% MEDs x Pix-In (Qtd)", "% Valor MEDs x Pix-In (Dia)", "Ação"
                ]]
                df_diario_unico = pd.concat([df_diario_unico, df_filtrado], ignore_index=True)

        with pd.ExcelWriter(saida_diaria, engine='openpyxl', mode='w') as writer_diaria:
            df_diario_unico.to_excel(writer_diaria, sheet_name="Análise Diária", index=False)

        formatar_cabecalhos(saida_diaria)
        print(f"\n📄 Relatório diário gerado com sucesso: {saida_diaria}")
    else:
        print("⚠️ Nenhuma empresa com dados no intervalo selecionado. Sem necessidade de gerar novo relatório.")

    formatar_cabecalhos(saida_geral)

